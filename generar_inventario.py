import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================
# DATOS TECNOLOGÍA (TECNO)
# =========================
data_tecno = [
    # Notebooks
    ["TEC-001", "Notebook Dell Latitude", "Dell", "Latitude 5430", "Notebook 14 pulgadas empresarial", "Tecnología", "Notebook", "Nuevo", "Compra Directa", 48, "SN-DEL5430001", "Intel Core i5-1235U", "16GB DDR4", "512GB SSD NVMe", "192.168.1.101", "Windows 11 Pro", "LAT-5430-001"],
    ["TEC-002", "Notebook HP ProBook", "HP", "ProBook 440 G9", "Notebook empresarial 14 pulgadas", "Tecnología", "Notebook", "Bueno", "Compra Directa", 48, "SN-HP440G9002", "Intel Core i7-1255U", "16GB DDR4", "512GB SSD NVMe", "192.168.1.102", "Windows 11 Pro", "HP-440G9-002"],
    ["TEC-003", "Notebook Lenovo ThinkPad", "Lenovo", "ThinkPad E14 Gen 4", "Notebook robusta para oficina", "Tecnología", "Notebook", "Nuevo", "Licitación", 48, "SN-LEN-E14-003", "AMD Ryzen 5 5625U", "8GB DDR4", "256GB SSD", "192.168.1.103", "Windows 11 Pro", "LEN-E14-003"],
    ["TEC-004", "Notebook ASUS VivoBook", "ASUS", "VivoBook 15", "Notebook uso general", "Tecnología", "Notebook", "Regular", "Donación", 36, "SN-ASU-VIV-004", "Intel Core i3-1115G4", "8GB DDR4", "256GB SSD", "192.168.1.104", "Windows 10 Home", "ASU-VIV-004"],
    
    # Monitores
    ["TEC-005", "Monitor Dell 24 pulgadas", "Dell", "P2422H", "Monitor Full HD IPS", "Tecnología", "Monitor", "Nuevo", "Compra Directa", 36, "SN-DEL-P24-005", None, None, None, None, None, None],
    ["TEC-006", "Monitor Samsung 27 pulgadas", "Samsung", "S27A600", "Monitor QHD profesional", "Tecnología", "Monitor", "Nuevo", "Compra Directa", 36, "SN-SAM-S27-006", None, None, None, None, None, None],
    ["TEC-007", "Monitor LG 22 pulgadas", "LG", "22MK430H", "Monitor Full HD uso general", "Tecnología", "Monitor", "Bueno", "Licitación", 36, "SN-LG-22M-007", None, None, None, None, None, None],
    ["TEC-008", "Monitor AOC 24 pulgadas", "AOC", "24B2XH", "Monitor económico Full HD", "Tecnología", "Monitor", "Regular", "Compra Directa", 24, "SN-AOC-24B-008", None, None, None, None, None, None],
    
    # Impresoras
    ["TEC-009", "Impresora HP LaserJet", "HP", "LaserJet Pro M404n", "Impresora láser monocromo red", "Tecnología", "Impresora", "Nuevo", "Compra Directa", 60, "SN-HP-M404-009", "600 MHz", "256MB", None, "192.168.1.201", "HP FutureSmart 4.0", "HP-M404-009"],
    ["TEC-010", "Impresora Canon multifunción", "Canon", "imageCLASS MF445dw", "Multifunción láser WiFi", "Tecnología", "Impresora", "Nuevo", "Licitación", 60, "SN-CAN-MF44-010", "800 MHz", "512MB", None, "192.168.1.202", "Canon OS", "CAN-MF44-010"],
    ["TEC-011", "Impresora Epson EcoTank", "Epson", "EcoTank L5290", "Multifunción tinta continua WiFi", "Tecnología", "Impresora", "Bueno", "Compra Directa", 48, "SN-EPS-L529-011", None, None, None, "192.168.1.203", None, "EPS-L529-011"],
    
    # Consumibles
    ["TEC-012", "Toner HP 58A negro", "HP", "CF258A", "Toner original LaserJet", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-HP58A-012", None, None, None, None, None, None],
    ["TEC-013", "Toner Canon 057 negro", "Canon", "3009C001", "Toner original imageCLASS", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-CAN057-013", None, None, None, None, None, None],
    ["TEC-014", "Botella tinta Epson 544 cian", "Epson", "T544", "Tinta original EcoTank cian", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-EPS544-014", None, None, None, None, None, None],
    ["TEC-015", "Botella tinta Epson 544 magenta", "Epson", "T544", "Tinta original EcoTank magenta", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-EPS544-015", None, None, None, None, None, None],
    ["TEC-016", "Botella tinta Epson 544 amarillo", "Epson", "T544", "Tinta original EcoTank amarillo", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-EPS544-016", None, None, None, None, None, None],
    ["TEC-017", "Botella tinta Epson 544 negro", "Epson", "T544", "Tinta original EcoTank negro", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-EPS544-017", None, None, None, None, None, None],
    ["TEC-018", "Papel bond A4 resma", "MultiMarcas", "A4-75gr", "Resma 500 hojas bond", "Tecnología", "Consumible", "Nuevo", "Compra Directa", None, "LOT-A4-75G-018", None, None, None, None, None, None],
    
    # Periféricos
    ["TEC-019", "Mouse Logitech inalámbrico", "Logitech", "M720 Triathlon", "Mouse ergonómico multisuperficie", "Tecnología", "Periférico", "Nuevo", "Compra Directa", 24, "SN-LOG-M720-019", None, None, None, None, None, None],
    ["TEC-020", "Mouse Microsoft ergonómico", "Microsoft", "Sculpt Ergonomic", "Mouse ergonómico vertical", "Tecnología", "Periférico", "Nuevo", "Compra Directa", 24, "SN-MS-SCUL-020", None, None, None, None, None, None],
    ["TEC-021", "Mouse Genius óptico", "Genius", "DX-120", "Mouse óptico básico USB", "Tecnología", "Periférico", "Bueno", "Donación", 12, "SN-GEN-DX12-021", None, None, None, None, None, None],
    ["TEC-022", "Teclado Logitech mecánico", "Logitech", "G413 SE", "Teclado mecánico gaming switches", "Tecnología", "Periférico", "Nuevo", "Compra Directa", 36, "SN-LOG-G413-022", None, None, None, None, None, None],
    ["TEC-023", "Teclado HP empresarial", "HP", "K1500", "Teclado USB estándar oficina", "Tecnología", "Periférico", "Nuevo", "Licitación", 24, "SN-HP-K150-023", None, None, None, None, None, None],
    ["TEC-024", "Teclado Dell silencioso", "Dell", "KB522", "Teclado USB con reposamuñecas", "Tecnología", "Periférico", "Bueno", "Compra Directa", 24, "SN-DEL-KB52-024", None, None, None, None, None, None],
    
    # Desktops
    ["TEC-025", "PC Dell OptiPlex", "Dell", "OptiPlex 7000 SFF", "Desktop compacto empresarial", "Tecnología", "Desktop", "Nuevo", "Compra Directa", 60, "SN-DEL-7000-025", "Intel Core i5-12500", "16GB DDR4", "512GB SSD NVMe", "192.168.1.105", "Windows 11 Pro", "DEL-7000-025"],
    ["TEC-026", "PC HP ProDesk", "HP", "ProDesk 400 G7", "Desktop microtorre oficina", "Tecnología", "Desktop", "Nuevo", "Licitación", 60, "SN-HP-400G7-026", "Intel Core i5-10500", "8GB DDR4", "256GB SSD", "192.168.1.106", "Windows 10 Pro", "HP-400G7-026"],
    
    # Redes
    ["TEC-027", "Switch TP-Link 24 puertos", "TP-Link", "TL-SG1024D", "Switch Gigabit rackmount", "Tecnología", "Redes", "Nuevo", "Compra Directa", 60, "SN-TP-SG24-027", None, None, None, "192.168.1.254", None, "TP-SG24-027"],
    ["TEC-028", "Router Cisco empresarial", "Cisco", "RV340", "Router VPN dual WAN", "Tecnología", "Redes", "Nuevo", "Compra Directa", 60, "SN-CIS-RV34-028", "Dual-core 900MHz", "512MB", "4GB Flash", "192.168.1.1", "Cisco IOS", "CIS-RV34-028"],
    ["TEC-029", "Access Point Ubiquiti", "Ubiquiti", "UAP-AC-LR", "Access Point largo alcance", "Tecnología", "Redes", "Nuevo", "Compra Directa", 48, "SN-UBI-ACLR-029", "QCA9563", "128MB", None, "192.168.1.253", "UniFi OS", "UBI-ACLR-029"],
    
    # Tablets y Audiovisual
    ["TEC-030", "Tablet Samsung Galaxy Tab", "Samsung", "Galaxy Tab A8", "Tablet 10.5 pulgadas WiFi", "Tecnología", "Tablet", "Nuevo", "Compra Directa", 36, "SN-SAM-A8-030", "Unisoc T618", "4GB", "64GB", "192.168.1.107", "Android 12", "SAM-A8-030"],
    ["TEC-031", "Webcam Logitech HD", "Logitech", "C920e", "Webcam Full HD empresarial", "Tecnología", "Audiovisual", "Nuevo", "Compra Directa", 24, "SN-LOG-C920-031", None, None, None, None, None, None],
    ["TEC-032", "Proyector Epson PowerLite", "Epson", "PowerLite E20", "Proyector XGA 3400 lúmenes", "Tecnología", "Audiovisual", "Nuevo", "Licitación", 60, "SN-EPS-E20-032", None, None, None, "192.168.1.204", None, "EPS-E20-032"],
    
    # Respaldo Energía
    ["TEC-033", "UPS APC 1500VA", "APC", "BR1500G", "UPS línea interactiva 8 tomas", "Tecnología", "Respaldo Energía", "Nuevo", "Compra Directa", 48, "SN-APC-1500-033", None, None, None, None, None, None],
    ["TEC-034", "Estabilizador Forza 2000VA", "Forza", "FVR-2002", "Estabilizador de voltaje 6 tomas", "Tecnología", "Respaldo Energía", "Nuevo", "Compra Directa", 36, "SN-FOR-2002-034", None, None, None, None, None, None],
]

# =========================
# DATOS MOBILIARIO (MUEBLE)
# =========================
data_mueble = [
    # Sillas
    ["MUB-001", "Silla ejecutiva ergonómica", "Braza", "Ergo Plus", "Silla respaldo alto con apoyabrazos ajustables", "Mobiliario", "Silla Ejecutiva", "Nuevo", "Compra Directa", 60, None, None, None, None, None, None, None],
    ["MUB-002", "Silla operativa estándar", "Novus", "Operativa 300", "Silla giratoria sin apoyabrazos", "Mobiliario", "Silla Operativa", "Nuevo", "Licitación", 48, None, None, None, None, None, None, None],
    ["MUB-003", "Silla visitante fija", "Maderkit", "Visit 100", "Silla fija patas metálicas", "Mobiliario", "Silla Visitante", "Nuevo", "Compra Directa", 48, None, None, None, None, None, None, None],
    ["MUB-004", "Silla gerencial cuero", "Braza", "President", "Silla gerencial cuero genuino base cromada", "Mobiliario", "Silla Gerencial", "Nuevo", "Compra Directa", 72, None, None, None, None, None, None, None],
    ["MUB-005", "Silla secretarial básica", "Genérica", "Secre 200", "Silla plástico respaldo bajo", "Mobiliario", "Silla Secretarial", "Bueno", "Donación", 36, None, None, None, None, None, None, None],
    ["MUB-006", "Silla auditorio apilable", "Maderkit", "Audit 50", "Silla apilable pata cromada", "Mobiliario", "Silla Auditorio", "Nuevo", "Compra Directa", 48, None, None, None, None, None, None, None],
    ["MUB-007", "Silla cafetería plástico", "Novus", "Cafe 100", "Silla plástico apilable color", "Mobiliario", "Silla Cafetería", "Nuevo", "Compra Directa", 36, None, None, None, None, None, None, None],
    
    # Escritorios
    ["MUB-008", "Escritorio ejecutivo L", "Maderkit", "Ejec L-180", "Escritorio en L 180x150cm cajonera", "Mobiliario", "Escritorio Ejecutivo", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-009", "Escritorio operativo recto", "Novus", "Oper 140", "Escritorio recto 140x60cm", "Mobiliario", "Escritorio Operativo", "Nuevo", "Licitación", 96, None, None, None, None, None, None, None],
    ["MUB-010", "Escritorio operativo recto", "Novus", "Oper 160", "Escritorio recto 160x70cm", "Mobiliario", "Escritorio Operativo", "Nuevo", "Licitación", 96, None, None, None, None, None, None, None],
    ["MUB-011", "Escritorio esquinero", "Maderkit", "Esqui 120", "Escritorio esquina 120x120cm", "Mobiliario", "Escritorio Esquinero", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-012", "Escritorio recepción", "Maderkit", "Recep 200", "Mostrador recepción 200x80cm", "Mobiliario", "Escritorio Recepción", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-013", "Mesa de juntas 8 personas", "Braza", "Juntas 240", "Mesa ovalada 240x120cm", "Mobiliario", "Mesa Juntas", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-014", "Mesa de juntas 12 personas", "Braza", "Juntas 360", "Mesa rectangular 360x140cm", "Mobiliario", "Mesa Juntas", "Nuevo", "Licitación", 120, None, None, None, None, None, None, None],
    ["MUB-015", "Escritorio operativo usado", "Genérica", "Oper 120", "Escritorio recto 120x60cm", "Mobiliario", "Escritorio Operativo", "Regular", "Donación", 48, None, None, None, None, None, None, None],
    
    # Almacenamiento
    ["MUB-016", "Cajonera móvil 3 cajones", "Novus", "Cajon 3C", "Cajonera bajo escritorio con ruedas", "Mobiliario", "Cajonera", "Nuevo", "Compra Directa", 72, None, None, None, None, None, None, None],
    ["MUB-017", "Archivero 4 gavetas", "Maderkit", "Archi 4G", "Archivero metálico 4 gavetas carta", "Mobiliario", "Archivero", "Nuevo", "Compra Directa", 96, None, None, None, None, None, None, None],
    ["MUB-018", "Archivero 2 gavetas", "Maderkit", "Archi 2G", "Archivero metálico 2 gavetas oficio", "Mobiliario", "Archivero", "Nuevo", "Licitación", 96, None, None, None, None, None, None, None],
    ["MUB-019", "Librero abierto 5 niveles", "Novus", "Libro 5N", "Librero melamina 5 niveles 180cm alto", "Mobiliario", "Librero", "Nuevo", "Compra Directa", 96, None, None, None, None, None, None, None],
    ["MUB-020", "Armario alto 2 puertas", "Maderkit", "Armar 2P", "Armario melamina 200x90x45cm", "Mobiliario", "Armario", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-021", "Armario bajo 1 puerta", "Maderkit", "Armar 1P", "Armado bajo 90x90x45cm", "Mobiliario", "Armario", "Bueno", "Donación", 72, None, None, None, None, None, None, None],
    
    # Complementario
    ["MUB-022", "Sofá recepción 3 cuerpos", "Braza", "Sofá 3C", "Sofá tela 3 cuerpos color gris", "Mobiliario", "Sofá", "Nuevo", "Compra Directa", 96, None, None, None, None, None, None, None],
    ["MUB-023", "Sofá recepción 2 cuerpos", "Braza", "Sofá 2C", "Sofá tela 2 cuerpos color azul", "Mobiliario", "Sofá", "Nuevo", "Compra Directa", 96, None, None, None, None, None, None, None],
    ["MUB-024", "Mesa centro recepción", "Maderkit", "Centro 80", "Mesa centro 80x80cm vidrio templado", "Mobiliario", "Mesa Centro", "Nuevo", "Compra Directa", 72, None, None, None, None, None, None, None],
    ["MUB-025", "Perchero de pie", "Novus", "Perchero P", "Perchero metal 12 ganchos", "Mobiliario", "Perchero", "Nuevo", "Compra Directa", 48, None, None, None, None, None, None, None],
    ["MUB-026", "Papelera metálica", "Genérica", "Papel M", "Papelera metálica 20 litros", "Mobiliario", "Papelera", "Nuevo", "Compra Directa", 36, None, None, None, None, None, None, None],
    ["MUB-027", "Pizarrón acrílico 120x90", "Novus", "Pizar 120", "Pizarrón blanco magnético", "Mobiliario", "Pizarrón", "Nuevo", "Compra Directa", 72, None, None, None, None, None, None, None],
    ["MUB-028", "Pizarrón acrílico 200x100", "Novus", "Pizar 200", "Pizarrón blanco magnético grande", "Mobiliario", "Pizarrón", "Nuevo", "Licitación", 72, None, None, None, None, None, None, None],
    ["MUB-029", "Biombo divisor 3 hojas", "Maderkit", "Biom 3H", "Biombo tela divisor ambientes", "Mobiliario", "Biombo", "Nuevo", "Compra Directa", 48, None, None, None, None, None, None, None],
    ["MUB-030", "Carrito de té", "Braza", "Té 2N", "Carrito servicio 2 niveles con ruedas", "Mobiliario", "Carrito Servicio", "Nuevo", "Compra Directa", 48, None, None, None, None, None, None, None],
    
    # Exterior
    ["MUB-031", "Banca plaza madera", "Maderkit", "Banca Ext", "Banca madera tratada 2m", "Mobiliario", "Banca Exterior", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-032", "Toldo retráctil 4x3m", "Genérica", "Toldo 43", "Toldo lona retráctil manual", "Mobiliario", "Toldo", "Nuevo", "Compra Directa", 96, None, None, None, None, None, None, None],
    ["MUB-033", "Cenicero exterior pie", "Novus", "Cen Ext", "Cenicero metal pie alto", "Mobiliario", "Cenicero", "Nuevo", "Compra Directa", 36, None, None, None, None, None, None, None],
    
    # Especial
    ["MUB-034", "Mesa dibujo técnico", "Maderkit", "Dibujo A0", "Mesa inclinable dibujo A0", "Mobiliario", "Mesa Especial", "Nuevo", "Compra Directa", 120, None, None, None, None, None, None, None],
    ["MUB-035", "Cabina telefónica insonorizada", "Braza", "Cabina Tel", "Cabina privacidad llamadas", "Mobiliario", "Cabina", "Nuevo", "Compra Directa", 96, None, None, None, None, None, None, None],
]

# =========================
# COMBINAR Y CREAR DATAFRAME
# =========================
all_data = data_tecno + data_mueble

columns = [
    "codigo_interno", "nombre", "marca", "modelo", "descripcion", 
    "categoria", "subcategoria", "condicion_fisica", "adquisicion", "vida_util_meses",
    "serial", "procesador", "memoria_ram", "disco_duro", "direccion_ip", "sistema_operativo", "host_name"
]

df = pd.DataFrame(all_data, columns=columns)

# =========================
# CREAR EXCEL CON FORMATO
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Inventario Municipal"

# Estilos
header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
tecno_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
mueble_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
border = Border(
    left=Side(style='thin', color="BDBDBD"),
    right=Side(style='thin', color="BDBDBD"),
    top=Side(style='thin', color="BDBDBD"),
    bottom=Side(style='thin', color="BDBDBD")
)

# Escribir datos
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            categoria = row[5] if len(row) > 5 else None
            if categoria == "Tecnología":
                cell.fill = tecno_fill
            elif categoria == "Mobiliario":
                cell.fill = mueble_fill

# Ajustar anchos
column_widths = {
    'A': 14, 'B': 32, 'C': 16, 'D': 22, 'E': 38,
    'F': 14, 'G': 22, 'H': 14, 'I': 18, 'J': 16,
    'K': 18, 'L': 20, 'M': 14, 'N': 18, 'O': 16,
    'P': 18, 'Q': 18
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Congelar paneles
ws.freeze_panes = 'A2'

# =========================
# HOJA DE RESUMEN
# =========================
ws_resumen = wb.create_sheet("Resumen")
ws_resumen['A1'] = "RESUMEN DE INVENTARIO MUNICIPAL"
ws_resumen['A1'].font = Font(bold=True, size=16, color="1B5E20")
ws_resumen.merge_cells('A1:B1')

resumen_data = [
    ["", ""],
    ["TOTALES GENERALES", ""],
    ["Items Tecnología", len(data_tecno)],
    ["Items Mobiliario", len(data_mueble)],
    ["TOTAL GENERAL", len(all_data)],
    ["", ""],
    ["DESGLOSE TECNOLOGÍA", ""],
    ["Notebooks", 4],
    ["Monitores", 4],
    ["Impresoras", 3],
    ["Consumibles", 7],
    ["Periféricos", 6],
    ["Desktops", 2],
    ["Redes", 3],
    ["Tablets", 1],
    ["Audiovisual", 2],
    ["Respaldo Energía", 2],
    ["", ""],
    ["DESGLOSE MOBILIARIO", ""],
    ["Sillas", 7],
    ["Escritorios", 8],
    ["Cajoneras / Archiveros / Libreros / Armarios", 6],
    ["Mobiliario Complementario", 9],
    ["Mobiliario Exterior", 3],
    ["Mobiliario Especial", 2],
]

for idx, (label, value) in enumerate(resumen_data, 3):
    cell_a = ws_resumen.cell(row=idx, column=1, value=label)
    cell_b = ws_resumen.cell(row=idx, column=2, value=value)
    
    if label in ["TOTALES GENERALES", "DESGLOSE TECNOLOGÍA", "DESGLOSE MOBILIARIO"]:
        cell_a.font = Font(bold=True, size=12, color="1B5E20")
    elif label and not label.startswith("-"):
        cell_a.font = Font(bold=True)
    
    if label.startswith("TOTAL"):
        cell_a.font = Font(bold=True, size=12, color="D32F2F")
        cell_b.font = Font(bold=True, size=12, color="D32F2F")

ws_resumen.column_dimensions['A'].width = 45
ws_resumen.column_dimensions['B'].width = 15

# =========================
# GUARDAR
# =========================
output_path = "Plantilla_Inventario_Pichidegua.xlsx"
wb.save(output_path)

print(f"✅ Excel generado exitosamente: {output_path}")
print(f"📊 Total items: {len(all_data)}")
print(f"   • Tecnología: {len(data_tecno)} items")
print(f"   • Mobiliario: {len(data_mueble)} items")
print(f"\n📋 Subcategorías incluidas:")
print(f"   TECNO: Notebooks, Monitores, Impresoras, Consumibles, Periféricos, Desktops, Redes, Tablets, Audiovisual, Respaldo Energía")
print(f"   MUEBLE: Sillas, Escritorios, Cajoneras, Archiveros, Libreros, Armarios, Sofás, Mesas, Pizarrones, Biombos, Exterior, Especial")
